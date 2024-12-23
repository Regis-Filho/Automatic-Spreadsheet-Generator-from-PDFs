import customtkinter
from tkinter import filedialog as fl
import openpyxl
from datetime import datetime
import datetime as dt
from PyPDF2 import PdfReader
from time import sleep


def selecionar_caminho():
    global caminhos
    caminhos = fl.askopenfilenames()


def gerar_palhilha():
    try:
        for caminho in caminhos:
            workbook = openpyxl.Workbook()
            validador = True
            # abrindo pdf
            with open(caminho, 'rb') as pdf_file:
                pdf_reader = PdfReader(pdf_file)

                # selecionando a pag
                page_coferir = pdf_reader.pages[0]
                # extraindo texto da pagina
                pdf_conferir = page_coferir.extract_text()
                if 'CONSULTA FERIAS UTILIZADAS' in pdf_conferir:
                    workbook.create_sheet(title=f'pag1')
                    del workbook['Sheet']
                    wb = workbook['pag1']
                    wb.append(['RELATIVAS',
                               'PRAZO',
                               'SITUACAO',
                               'DOBRA',
                               'ABONO',
                               'QTD_DIAS_ABONO',
                               'G1INI',
                               'G1FIM',
                               'G1DOBRA',
                               'G2INI',
                               'G2FIM',
                               'G2DOBRA',
                               'G3INI',
                               'G3FIM',
                               'G3DOBRA'])

                    for i in range(len(pdf_reader.pages)):

                        print(f'{"*"*20}PAGINA {i+1} {"*"*20}')
                        # selecionando a pag
                        page = pdf_reader.pages[i]
                        # extraindo texto da pagina
                        pdf = page.extract_text()

                        if 'Documento assinado' in pdf:
                            index_final = pdf.index('Documento assinado')
                        
                        else:
                            index_final = pdf.index('Assinado')

                        index_comeco = pdf.index('Data Inicio')

                        pdf = pdf[index_comeco:index_final-1]
                        try:
                            index_numero_processo = pdf.index(
                                'Número do processo:')

                            tabela = pdf[:index_numero_processo-1].split('\n')

                        except:
                            tabela = pdf.split('\n')

                        tabela.pop(0)

                        linhas = [linha.split()for linha in tabela]

                        for linha in linhas:
                            linha.remove(linha[-1])

                        linhas_formatada = []

                        for linha in linhas:
                            print(linha)
                            if '/' in linha[0] and '/' in linha[1] and '/' in linha[2] and '/' in linha[3]:
                                linhas_formatada.append(linha)

                            elif '/' in linha[0] and '/' in linha[1] and '/' in linha[2]:
                                linhas_formatada.append(linha)

                            else:
                                for palavra in linha:
                                    linhas_formatada[-1].append(palavra)

                        for linha in linhas_formatada:
                            if len(linha) == 6:
                                ano1 = linha[0].split('/')[-1]
                                ano2 = linha[1].split('/')[-1]
                                ano1_ano2 = f'{ano1}/{ano2}'

                                wb.append([
                                    ano1_ano2,
                                    '30',
                                    'G',
                                    'N',
                                    'S',
                                    linha[-1],
                                    linha[2],
                                    linha[3],
                                    'N',
                                    '',
                                    '',
                                    '',
                                    '',
                                    '',
                                    ''
                                ])
                            elif len(linha) == 5:
                                ano1 = linha[0].split('/')[-1]
                                ano2 = linha[1].split('/')[-1]
                                ano1_ano2 = f'{ano1}/{ano2}'

                                wb.append([
                                    ano1_ano2,
                                    '30'if linha[-1] == '30' else linha[-1],
                                    'G'if linha[-1] == '30' else 'GP',
                                    'N',
                                    'N',
                                    '0',
                                    linha[2],
                                    linha[3],
                                    'N',
                                    '',
                                    '',
                                    '',
                                    '',
                                    '',
                                    ''
                                ])
                            elif len(linha) == 8:
                                ano1 = linha[0].split('/')[-1]
                                ano2 = linha[1].split('/')[-1]
                                ano1_ano2 = f'{ano1}/{ano2}'

                                wb.append([
                                    ano1_ano2,
                                    '30',
                                    'G',
                                    'N',
                                    'N',
                                    '0',
                                    linha[5],
                                    linha[6],
                                    'N',
                                    linha[2],
                                    linha[3],
                                    'N',
                                    '',
                                    '',
                                    ''
                                ])

                            elif len(linha) == 9:
                                ano1 = linha[0].split('/')[-1]
                                ano2 = linha[1].split('/')[-1]
                                ano1_ano2 = f'{ano1}/{ano2}'
                                if '/' in linha[5]:
                                    wb.append([
                                        ano1_ano2,
                                        '30',
                                        'G',
                                        'N',
                                        'S',
                                        linha[-1],
                                        linha[5],
                                        linha[6],
                                        'N',
                                        linha[2],
                                        linha[3],
                                        'N',
                                        '',
                                        '',
                                        ''
                                    ])
                                else:
                                    wb.append([
                                        ano1_ano2,
                                        '30',
                                        'G',
                                        'N',
                                        'S',
                                        linha[5],
                                        linha[6],
                                        linha[7],
                                        'N',
                                        linha[2],
                                        linha[3],
                                        'N',
                                        '',
                                        '',
                                        ''
                                    ])
                            elif len(linha) == 4:
                                if not '/' in linha[-1] and '/' in linha[-2]:
                                    ano1 = linha[0].split('/')[-1]
                                    ano2 = linha[1].split('/')[-1]
                                    ano1_ano2 = f'{ano1}/{ano2}'

                                    wb.append([
                                        ano1_ano2,
                                        '30'if linha[-1] == '30' else linha[-1],
                                        'G'if linha[-1] == '30' else 'P',
                                        'N',
                                        'N',
                                        '0',
                                        linha[2],
                                        '',
                                        'N',
                                        '',
                                        '',
                                        '',
                                        '',
                                        '',
                                        ''
                                    ])

                            elif len(linha) == 10:
                                ano1 = linha[0].split('/')[-1]
                                ano2 = linha[1].split('/')[-1]
                                ano1_ano2 = f'{ano1}/{ano2}'

                                wb.append([
                                    ano1_ano2,
                                    '30',
                                    'G',
                                    'N',
                                    'S',
                                    f'{int(linha[5])+int(linha[-1])}',
                                    linha[6],
                                    linha[7],
                                    'N',
                                    linha[2],
                                    linha[3],
                                    'N',
                                    '',
                                    '',
                                    ''
                                ])
                            elif len(linha) == 11:
                                ano1 = linha[0].split('/')[-1]
                                ano2 = linha[1].split('/')[-1]
                                ano1_ano2 = f'{ano1}/{ano2}'

                                wb.append([
                                    ano1_ano2,
                                    '30',
                                    'G',
                                    'N',
                                    'N',
                                    f'0',
                                    linha[8],
                                    linha[9],
                                    'N',
                                    linha[5],
                                    linha[6],
                                    'N',
                                    linha[2],
                                    linha[3],
                                    'N'
                                ])
                        sleep(1)
                        data_horario = str(datetime.today().strftime(
                            '%d/%m/%Y--%H:%M:%S')).replace('/', '-').replace(':', '-')

                        workbook.save(f'planilha_Ferias_{data_horario}.xlsx')
                        sleep(1)
                        workbook = None

                elif 'CONSULTA FINANCEIRO MENSAL' in pdf_conferir:
                    workbook['Sheet'].title = 'pag1'

                    workbook_pag1 = workbook['pag1']

                    workbook_pag1.append(['DATA',
                                          'SALARIO',
                                          'FUNCAO GRATIFICADA EFETIVA',
                                          'FUNCAO GRATIFICADA NAO EFET',
                                          'AC  FUNCAO GRATIFICADA NAO',
                                          'CTVA - FG/CC NAO EFETIV',
                                          'SERVICO EXTRAORDINARIO',
                                          'MEDIA HORA EXTRA',
                                          'MEDIA HORA EXTRA FERIAS',
                                          'REMUNERACAO DE FERIAS',
                                          'REMUNERACAO 1/3 DAS FERIAS',
                                          'MEDIA HE/ADIC SOBR FERIAS',
                                          'ADIANTAMENTO DE SALARIO DE',
                                          'ABONO PECUNIARIO',
                                          '1/3 SOBRE ABONO PECUNIARIO',
                                          'PLR FENABAN',
                                          'PLR CAIXA',
                                          'REMUNERACAO BASE',
                                          'BASE INSS EMPREG'])
                    pdf_reader = PdfReader(pdf_file)
                    for i in range(len(pdf_reader.pages)):
                        print(f'{"*"*20}PAGINA {i+1} {"*"*20}')
                        page = pdf_reader.pages[i]

                        pdf = page.extract_text()

                        linhas = pdf.split('\n')

                        data = None
                        salario = 0
                        salario_c = 0
                        ac_salario = 0

                        gratificante_01 = 0
                        gratificacao_02 = 0

                        funcao_gratificada_nao_efetiva = 0
                        ac_funcao_gratificada_nao = 0

                        nao_efeiva_CTVA = 0

                        servico_extraordinario = 0

                        media_hora_extra = 0
                        media_hora_extra_ferias = 0

                        remuneracao_ferias = 0
                        remuneracao_1_3_das_ferias = 0

                        media_he_adic = 0
                        media_he_adic_sobr_ferias = 0

                        adiantamento_salario_de = 0

                        abono_pecuniario = 0
                        abono_pecuniario_1_3 = 0

                        plr_fenaban = 0
                        plr_caixa = 0

                        remuneracao_base = 0
                        base_inss_empreg = 0

                        data = f'{linhas[4].split()[2]}/{linhas[4].split()[4]}'
                        print(data)

                        for linha in linhas:

                            # salario
                            if '0002 01 C SALARIO PADRAO' in linha and 'SALARIO' in linha:

                                salario = float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(f'salario padrao: {salario}')

                            if '0002 02 C SALARIO PADRAO ' in linha and 'SALARIO' in linha:

                                salario_c = float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(f'salario padrao: {salario_c}')

                            if '0002 01 C AC  SALARIO PADRAO' in linha and 'SALARIO' in linha:

                                ac_salario = float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(f'salario padrao: {ac_salario}')

                            # 01 C FUNCAO GRATIFICADA EFETIVA
                            if '01 C FUNCAO GRATIFICADA EFETIVA' in linha:
                                gratificante_01 = float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f'01 C FUNCAO GRATIFICADA EFETIVA :{gratificante_01}')

                            # 0275 02 C FUNCAO GRATIFICADA EFETIVA
                            if '0275 02 C FUNCAO GRATIFICADA EFETIVA' in linha:
                                gratificacao_02 = float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f'0275 02 C FUNCAO GRATIFICADA EFETIVA : {gratificacao_02}')

                            # 10 C FUNCAO GRATIFICADA NAO EFET
                            if '10 C FUNCAO GRATIFICADA NAO EFET' in linha:
                                funcao_gratificada_nao_efetiva = float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f'001-Adicional Tempo Serviço : {funcao_gratificada_nao_efetiva}')

                            #  AC  FUNCAO GRATIFICADA NAO
                            if ' AC  FUNCAO GRATIFICADA NAO ' in linha:
                                ac_funcao_gratificada_nao = float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' AC  FUNCAO GRATIFICADA NAO  : {ac_funcao_gratificada_nao}')

                            # CTVA - FG/CC NAO EFETIVA
                            if 'CTVA - FG/CC NAO EFETIVA' in linha:
                                nao_efeiva_CTVA = float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f'CTVA - FG/CC NAO EFETIVA:{nao_efeiva_CTVA}')

                            # 2 0044 01 C SERVICO EXTRAORDINARIO
                            if '2 0044 00 C SERVICO EXTRAORDINARIO  ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0044 01 C SERVICO EXTRAORDINARIO  :{servico_extraordinario}')

                            # 2 0044 01 C SERVICO EXTRAORDINARIO
                            if '2 0044 01 C SERVICO EXTRAORDINARIO  ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0044 01 C SERVICO EXTRAORDINARIO  :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 02 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 03 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 04 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 05 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 06 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 07 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 08 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 09 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')
                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 10 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')

                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 11 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')

                            # 2 0044 11 C SERVICO EXTRAORDINARIO
                            if ' 2 0044 12 C SERVICO EXTRAORDINARIO ' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0044 11 C SERVICO EXTRAORDINARIO :{servico_extraordinario}')

                            # 2 0244 01 C SERVICO EXTRAORDINARIO AUTO
                            if ' 2 0244 01 C SERVICO EXTRAORDINARIO AUTO' in linha:
                                servico_extraordinario += float(
                                    linha.split()[-2].replace('.', '').replace(',', '.'))
                                print(
                                    f' 2 0244 01 C SERVICO EXTRAORDINARIO AUTO:{servico_extraordinario}')

                            # 2 0057 01 C MEDIA HORA EXTRA - REPOUSO
                            if '2 0057 01 C MEDIA HORA EXTRA - REPOUSO' in linha:
                                media_hora_extra += float(linha.split()
                                                          [-2].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0057 01 C MEDIA HORA EXTRA - REPOUSO : {media_hora_extra}')

                            # 2 0058 01 C MEDIA HORA EXTRA AUTORIZADA
                            if '2 0058 01 C MEDIA HORA EXTRA AUTORIZADA' in linha:
                                media_hora_extra += float(linha.split()
                                                          [-2].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0058 01 C MEDIA HORA EXTRA AUTORIZADA : {media_hora_extra}')

                            # 2 1145 04 C MEDIA HORA EXTRA FERIAS MES
                            if '2 1145 04 C MEDIA HORA EXTRA FERIAS MES' in linha:
                                media_hora_extra_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1145 04 C MEDIA HORA EXTRA FERIAS MES : {media_hora_extra_ferias}')

                            # 2 1145 08 C MEDIA HORA EXTRA FERIAS MES
                            if '2 1145 08 C MEDIA HORA EXTRA FERIAS MES' in linha:
                                media_hora_extra_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1145 04 C MEDIA HORA EXTRA FERIAS MES : {media_hora_extra_ferias}')

                            # 2 1100 00 C REMUNERACAO DE FERIAS
                            if '2 1100 00 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 00 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 01 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 02 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 03 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 04 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 05 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')
                            try:
                                # 2 1100 01 C REMUNERACAO DE FERIAS
                                if '2 1100 06 C REMUNERACAO DE FERIAS' in linha:
                                    remuneracao_ferias += float(
                                        linha.split()[-3].replace('.', '').replace(',', '.'))
                                    print(
                                        f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')
                                    
                            except:
                                if '2 1100 06 C REMUNERACAO DE FERIAS' in linha:
                                    remuneracao_ferias += float(
                                        linha.split()[-2].replace('.', '').replace(',', '.'))
                                    print(
                                        f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 07 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 08 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 09 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 10 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 11 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1100 01 C REMUNERACAO DE FERIAS
                            if '2 1100 12 C REMUNERACAO DE FERIAS' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1100 01 C REMUNERACAO DE FERIAS: {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 00 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 01 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 02 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 03 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 04 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 05 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 06 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 07 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 08 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 09 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 10 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 11 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 1300 02 C REMUNERACAO FERIAS MES SEGU
                            if '2 1300 12 C REMUNERACAO FERIAS MES SEGU ' in linha:
                                remuneracao_ferias += float(
                                    linha.split()[-3].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1300 02 C REMUNERACAO FERIAS MES SEGU : {remuneracao_ferias}')

                            # 2 0043 00 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 00 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 00 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 01 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 02 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 03 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 04 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 05 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 06 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 07 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 08 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 09 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 10 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 11 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 01 C REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 12 C REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0043 10 I REMUNERACAO 1/3 DAS FERIAS
                            if '2 0043 00 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 01 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 02 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 03 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 04 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 05 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 06 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 07 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 08 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 09 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 10 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 11 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')
                            if '2 0043 12 I REMUNERACAO 1/3 DAS FERIAS' in linha:
                                remuneracao_1_3_das_ferias += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0043 01 C REMUNERACAO 1/3 DAS FERIAS : {remuneracao_1_3_das_ferias}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 00 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 01 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 02 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 03 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 04 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 05 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 06 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 07 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 08 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 09 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 10 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 11 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 12 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 00 C MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 12 C MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 0045 10 I MEDIA HE/ADIC SOBR FERIAS
                            if '2 0045 00 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 01 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 02 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 03 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 04 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 05 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 06 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 07 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 08 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 09 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 10 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 11 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')
                            if '2 0045 12 I MEDIA HE/ADIC SOBR FERIAS' in linha:
                                media_he_adic += float(linha.split()
                                                       [9].replace('.', '').replace(',', '.'))
                                print(
                                    f'MEDIA HE/ADIC SOBR FERIAS : {media_he_adic}')

                            # 2 1200 10 C ADIANTAMENTO DE SALARIO DE
                            if ' 2 1200 00 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 01 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 02 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 03 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 04 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 05 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 06 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 07 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 08 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 09 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 10 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 11 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if ' 2 1200 12 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')

                            # 2 1201 10 C ADIANTAMENTO DE SALARIO DE
                            if '2 1201 00 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 01 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 02 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 03 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 04 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 05 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 06 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 07 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 08 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 09 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 10 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 11 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')
                            if '2 1201 12 C ADIANTAMENTO DE SALARIO DE' in linha:
                                adiantamento_salario_de += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 1200 00 C ADIANTAMENTO DE SALARIO DE: {adiantamento_salario_de}')

                            # 2 0065 02 S ABONO PECUNIARIO
                            if '2 0065 00 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 01 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 02 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 03 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 04 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 05 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 06 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 07 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 08 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 09 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 10 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 11 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 12 S ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')

                            # 1 0065 03 C AC  ABONO PECUNIARIO
                            if '1 0065 00 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 01 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 02 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 03 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 04 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 05 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 06 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 07 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 08 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 09 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 10 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 11 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')
                            if '1 0065 12 C AC  ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [8].replace('.', '').replace(',', '.'))
                                print(
                                    f'1 0065 03 C AC  ABONO PECUNIARIO: {abono_pecuniario}')

                            # 2 0065 08 C ABONO PECUNIARIO
                            if '2 0065 00 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 01 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 02 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 03 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 04 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 05 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 06 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 07 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 08 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 09 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 10 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 11 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')
                            if '2 0065 12 C ABONO PECUNIARIO' in linha:
                                abono_pecuniario += float(linha.split()
                                                          [7].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0065 08 C ABONO PECUNIARIO: {abono_pecuniario}')

                            # 2 0208 08 C 1/3 SOBRE ABONO PECUNIARIO
                            if '2 0208 00 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 01 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 02 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 03 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 04 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 05 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 06 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 07 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 08 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 09 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 10 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 11 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 12 C 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')

                            # 2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO
                            if '2 0208 00 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 01 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 03 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 04 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 05 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 06 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 07 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 08 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 09 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 10 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 11 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')
                            if '2 0208 12 S 1/3 SOBRE ABONO PECUNIARIO' in linha:
                                abono_pecuniario_1_3 += float(
                                    linha.split()[9].replace('.', '').replace(',', '.'))
                                print(
                                    f'2 0208 02 S 1/3 SOBRE ABONO PECUNIARIO: {abono_pecuniario_1_3}')

                            # 2 2 0230 50 I PLR FENABAN
                            if '2 0230 50 I PLR FENABAN   ' in linha:
                                plr_fenaban += float(linha.split()
                                                     [7].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')
                            if '2 0230 30 I PLR FENABAN   ' in linha:
                                plr_fenaban += float(linha.split()
                                                     [7].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')

                            if '1 0230 30 I AC  PLR FENABAN  ' in linha:
                                plr_fenaban += float(linha.split()
                                                     [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')
                            if '1 0230 50 I AC  PLR FENABAN  ' in linha:
                                plr_fenaban += float(linha.split()
                                                     [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')

                            # 114 2 1232 50 I PLR FENABAN 2
                            if '2 1232 50 I PLR FENABAN 2' in linha:
                                plr_fenaban += float(linha.split()
                                                     [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')
                            if '2 1232 30 I PLR FENABAN 2' in linha:
                                plr_fenaban += float(linha.split()
                                                     [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')

                            # 114 3 0230 50 I REP PLR FENABAN
                            if '3 0230 50 I REP PLR FENABAN' in linha:
                                plr_fenaban += float(linha.split()
                                                     [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')
                            if '3 0230 30 I REP PLR FENABAN' in linha:
                                plr_fenaban += float(linha.split()
                                                     [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')

                            # 114 3 1232 50 I REP PLR FENABAN 2
                            if '3 1232 50 I REP PLR FENABAN 2' in linha:
                                plr_fenaban += float(linha.split()
                                                     [9].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')
                            if '3 1232 30 I REP PLR FENABAN 2' in linha:
                                plr_fenaban += float(linha.split()
                                                     [9].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_fenaban}')

                            # 099 2 0231 30 I PLR CAIXA
                            if '2 0231 30 I PLR CAIXA' in linha:
                                plr_caixa += float(linha.split()
                                                   [7].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_caixa}')
                            if '2 0231 50 I PLR CAIXA' in linha:
                                plr_caixa += float(linha.split()
                                                   [7].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_caixa}')
                            if '1 0231 30 I AC  PLR CAIXA' in linha:
                                plr_caixa += float(linha.split()
                                                   [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_caixa}')
                            if '1 0231 50 I AC  PLR CAIXA' in linha:
                                plr_caixa += float(linha.split()
                                                   [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_caixa}')

                            if '3 0231 50 I REP PLR CAIXA' in linha:
                                plr_caixa += float(linha.split()
                                                   [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_caixa}')
                            if '3 0231 30 I REP PLR CAIXA' in linha:
                                plr_caixa += float(linha.split()
                                                   [8].replace('.', '').replace(',', '.'))
                                print(f'PLR FENABAN: {plr_caixa}')

                            # Remuneracao Base :
                            if 'Remuneracao Base :' in linha:
                                remuneracao_base = float(
                                    linha.split()[3].replace('.', '').replace(',', '.'))
                                print(
                                    f'Remuneracao Base :: {remuneracao_base}')

                            # Base INSS Empreg.:
                            if 'Base INSS Empreg.:' in linha:
                                base_inss_empreg = float(
                                    linha.split()[3].replace('.', '').replace(',', '.'))
                                print(
                                    f'Remuneracao Base :: {base_inss_empreg}')

                        salario_final = salario+salario_c+ac_salario

                        gratificante_efetiva = gratificacao_02+gratificante_01

                        workbook_pag1.append([data,
                                              salario_final,
                                              gratificante_efetiva,
                                              funcao_gratificada_nao_efetiva,
                                              ac_funcao_gratificada_nao,
                                              nao_efeiva_CTVA,
                                              servico_extraordinario,
                                              media_hora_extra,
                                              media_hora_extra_ferias,
                                              remuneracao_ferias,
                                              remuneracao_1_3_das_ferias,
                                              media_he_adic,
                                              adiantamento_salario_de,
                                              abono_pecuniario,
                                              abono_pecuniario_1_3,
                                              plr_fenaban,
                                              plr_caixa,
                                              remuneracao_base,
                                              base_inss_empreg





                                              ])

                    sleep(1)
                    data_horario = str(datetime.today().strftime(
                        '%d/%m/%Y--%H:%M:%S')).replace('/', '-').replace(':', '-')

                    workbook.save(f'planilha_contra_cheque{data_horario}.xlsx')
                    sleep(1)

                elif 'CONSULTA PONTO ELETRONICO' in pdf_conferir:
                    workbook = openpyxl.Workbook()

                    def remover_item_repetido(lista):
                        contador = 0
                        elemento = []
                        for palavra in lista:
                            if palavra in elemento:
                                lista.pop(contador)

                            else:

                                elemento.append(palavra)
                                contador += 1

                    pdf_reader = PdfReader(pdf_file)

                    workbook.create_sheet(title=f'pag{1}')
                    del workbook['Sheet']
                    wb = workbook[f'pag1']

                    wb.append(['Data', 'Entrada1', 'SaÃ-da1', 'Entrada2', 'SaÃ-da2', 'Entrada3',
                              'SaÃ-da3', 'Entrada4', 'SaÃ-da4', 'Entrada5', 'SaÃ-da5', 'Entrada6', 'SaÃ-da6'])
                    # passando por cada pag do pdf
                    for i in range(len(pdf_reader.pages)):
                        try:
                            print(f'{"*"*20}PAGINA {i+1} {"*"*20}')
                            # selecionando a pag
                            page = pdf_reader.pages[i]
                            # extraindo texto da pagina
                            pdf = page.extract_text()

                            # pegando os index do começo e final da tabela
                            try:
                                try:

                                    # pegar mes/ano

                                    index_mes_ano = pdf.index('Mes/Ano')
                                    index_tipo_jornada = pdf.index(
                                        'Tipo de Jornada')
                                    linha_mes_ano = pdf[index_mes_ano:index_tipo_jornada-1]
                                    mes_ano = linha_mes_ano[linha_mes_ano.index(
                                        ':')+1:]

                                    mes_ano = mes_ano.replace(' ', '')

                                    index_comeco = pdf.index('Dia Semana')
                                    if 'Documento assinado' in pdf:
                                        index_final = pdf.index('Documento assinado')
                        
                                    else:
                                        index_final = pdf.index('Assinado')
                                    
                                    try:

                                        index_final_numero = pdf.index(
                                            'Número do processo')

                                    except:
                                        index_final_numero = 100000

                                # aplicando os index para pegar a tabela
                                    if index_final_numero < index_final:
                                        tabela = pdf[index_comeco:index_final_numero-1]

                                    else:
                                        tabela = pdf[index_comeco:index_final-1]

                                except:
                                    index_final = pdf.index('Assinado')
                                    tabela = pdf[:index_final-1]

                                # quebrando  a tabela em linhas
                                tabela = tabela.split('\n')

                                # removendo o cabeçalho da tabela
                                if 'Dia' in tabela[0]:
                                    tabela.pop(0)

                                # criando uma varivel para armazenar o conteudo extraido
                                arquivo_com_dia_repetido = []

                                # passando por cada linha na tabela
                                for linha in tabela:
                                    # trocando '-' por '' e separando as palavras e formando uma lista
                                    linha_extraida = linha.replace(
                                        '-', '').split()
                                    # selecionando apenas as linha que tem mais de 1 valor
                                    if len(linha_extraida) > 1:
                                        for i in range(2):
                                            # removendo as palavras que tem na lista
                                            for palavra in linha_extraida:
                                                try:
                                                    if ':' in palavra:
                                                        conteudo1 = int(
                                                            palavra.split(':')[0])

                                                    else:
                                                        conteudo1 = int(
                                                            palavra)

                                                except:
                                                    linha_extraida.remove(
                                                        palavra)

                                        # adicionando a linha extraida a lista arquivo
                                        arquivo_com_dia_repetido.append(
                                            linha_extraida)

                                horario_permitido = dt.time(hour=7, minute=25)

                                for linha in arquivo_com_dia_repetido:
                                    try:
                                        for palavra in linha[2:]:
                                            hora = int(palavra.split(':')[0])
                                            minuto = int(palavra.split(':')[1])
                                            quantidade = dt.time(
                                                hour=hora, minute=minuto)

                                            if quantidade < horario_permitido:
                                                linha.remove(palavra)

                                    except:
                                        pass

                                for lista in arquivo_com_dia_repetido:
                                    remover_item_repetido(lista)

                                for linha in arquivo_com_dia_repetido:
                                    if len(linha) > 2:
                                        for palavra in linha[2:]:
                                            if not ':' in palavra:
                                                linha.remove(palavra)

                                # removendo os valores que são a quantidades
                                horario_permitido = dt.time(hour=7, minute=25)

                                for linha in arquivo_com_dia_repetido:
                                    try:
                                        for palavra in linha[2:]:
                                            hora = int(palavra.split(':')[0])
                                            minuto = int(palavra.split(':')[1])
                                            quantidade = dt.time(
                                                hour=hora, minute=minuto)

                                            if quantidade < horario_permitido:
                                                linha.remove(palavra)

                                    except:
                                        pass

                                # remover listas vazia
                                for i in range(3):
                                    for lista in arquivo_com_dia_repetido:
                                        if len(lista) == 0:
                                            arquivo_com_dia_repetido.remove(
                                                lista)

                                # adicionando as quantidas faltantes
                                for lista in arquivo_com_dia_repetido:
                                    if len(lista) == 1:
                                        try:
                                            item = int(lista[0])
                                            lista.insert(1, '06:00')

                                        except:
                                            pass

                                #  remover dia repetido
                                # remover listas vazia
                                for i in range(3):
                                    for lista in arquivo_com_dia_repetido:
                                        if len(lista) == 1:
                                            arquivo_com_dia_repetido.remove(
                                                lista)

                                dia_repetido = None

                                for linha in arquivo_com_dia_repetido:
                                    dia = linha[0]
                                    jornada = linha[1]
                                    if dia == dia_repetido:

                                        linha.remove(dia)
                                        if len(linha) % 2 != 0:
                                            linha.pop(0)

                                    dia_repetido = dia

                                print(arquivo_com_dia_repetido)

                                # remover listas vazia e que tem 1 item
                                for i in range(3):
                                    for lista in arquivo_com_dia_repetido:
                                        if len(lista) == 0:
                                            arquivo_com_dia_repetido.remove(
                                                lista)

                                # for lista in arquivo_com_dia_repetido:
                                    if len(lista) == 3:
                                        if ':' in lista[0]:
                                            lista.remove(lista[0])

                                # juntando os horarios
                                for i in range(3):
                                    for linha in arquivo_com_dia_repetido:
                                        if len(linha) == 2:
                                            if int(linha[1][0:2]) > 8:
                                                palavras = linha
                                                index_palavra = arquivo_com_dia_repetido.index(
                                                    linha)
                                                arquivo_com_dia_repetido.pop(
                                                    index_palavra)
                                                for palavra in palavras:
                                                    arquivo_com_dia_repetido[index_palavra-1].append(
                                                        palavra)

                                # remover listas vazia e que tem 1 item
                                for i in range(3):
                                    for lista in arquivo_com_dia_repetido:
                                        if len(lista) == 1:
                                            arquivo_com_dia_repetido.remove(
                                                lista)

                                #  #  remover dia repetido

                                # print(arquivo_com_dia_repetido)
                                dia_repetido = None
                                for linha in arquivo_com_dia_repetido:
                                    dia = linha[0]
                                    if dia == dia_repetido:
                                        linha.remove(linha[0])
                                        linha.remove(linha[0])

                                    dia_repetido = dia

                                    # remover listas vazia e que tem 1 item
                                for i in range(3):
                                    for lista in arquivo_com_dia_repetido:
                                        if len(lista) == 1:
                                            arquivo_com_dia_repetido.remove(
                                                lista)

                                    # remover listas vazia e que tem 1 item
                                for i in range(3):
                                    for lista in arquivo_com_dia_repetido:
                                        if len(lista) == 0:
                                            arquivo_com_dia_repetido.remove(
                                                lista)

                                    # juntando os horarios
                                for i in range(3):
                                    for linha in arquivo_com_dia_repetido:
                                        if len(linha) == 2:
                                            if int(linha[1][0:2]) > 8:
                                                palavras = linha
                                                index_palavra = arquivo_com_dia_repetido.index(
                                                    linha)
                                                arquivo_com_dia_repetido.pop(
                                                    index_palavra)
                                                for palavra in palavras:
                                                    arquivo_com_dia_repetido[index_palavra-1].append(
                                                        palavra)

                                # horarios faltandes de pdfs com apenas 1 entrada
                                for lista in arquivo_com_dia_repetido:
                                    if len(lista) == 3 or len(lista) == 5 or len(lista) == 7:
                                        if int(lista[1].split(':')[0]) >= 7:
                                            lista.insert(1, '06:00')

                                for lista in arquivo_com_dia_repetido:
                                    if len(lista) == 3:
                                        # carga horaria
                                        carga_horaria_hora = int(
                                            lista[1].split(':')[0])
                                        carga_horaria_minuto = int(
                                            lista[1].split(':')[1])
                                        carga_horaria = dt.timedelta(
                                            hours=carga_horaria_hora, minutes=carga_horaria_minuto)

                                        # entrada 1
                                        entrada_1_hora = int(
                                            lista[2].split(':')[0])
                                        entrada_1_minuto = int(
                                            lista[2].split(':')[1])
                                        entrada_1 = dt.datetime.combine(
                                            dt.date.today(), dt.time(entrada_1_hora, entrada_1_minuto))

                                        # somando entrada com a carga horaria:
                                        resultado_saida_1_calculo = entrada_1+carga_horaria
                                        saida_2 = resultado_saida_1_calculo.strftime(
                                            '%H:%M')
                                        lista.append(saida_2)

                                for lista in arquivo_com_dia_repetido:
                                    if len(lista) == 2:
                                        wb.append(
                                            [f'{lista[0]}/{mes_ano}', '', '', '', '', '', '', '', ''])
                                    elif len(lista) == 4:
                                        wb.append(
                                            [f'{lista[0]}/{mes_ano}', lista[2], lista[3], '', '', '', '', '', '', '', '', '', ''])
                                    elif len(lista) == 6:
                                        wb.append([f'{lista[0]}/{mes_ano}', lista[2], lista[3],
                                                  lista[4], lista[5], '', '', '', '', '', '', '', ''])
                                    elif len(lista) == 8:
                                        wb.append([f'{lista[0]}/{mes_ano}', lista[2], lista[3], lista[4],
                                                  lista[5], lista[6], lista[7], '', '', '', '', '', ''])
                                    elif len(lista) == 10:
                                        wb.append([f'{lista[0]}/{mes_ano}', lista[2], lista[3], lista[4],
                                                  lista[5], lista[6], lista[7], lista[8], lista[9], '', '', '', ''])

                            except:
                                pass

                        except IndexError as erro:
                            print(erro)

                    sleep(1)
                    data_horario = str(datetime.today().strftime(
                        '%d/%m/%Y--%H:%M:%S')).replace('/', '-').replace(':', '-')
                    workbook.save(f'planilha_ponto_{data_horario}.xlsx')
                    workbook = None
                    sleep(1)

                else:
                    validador = False
                    janela_nao_encontrado = customtkinter.CTk()

                    janela_nao_encontrado.geometry('300x100')
                    janela_nao_encontrado.resizable(width=False, height=False)
                    janela_nao_encontrado.title('PDF NÃO ENCONTRADO')

                    texto_nao_encontrado = customtkinter.CTkLabel(
                        janela_nao_encontrado, text='PDF NÃO ENCONTRADO')
                    texto_nao_encontrado.pack(padx=10, pady=10)

                    botao_ok = customtkinter.CTkButton(
                        janela_nao_encontrado, text='ok', command=janela_nao_encontrado.destroy)
                    botao_ok.pack(padx=10, pady=10)

                    janela_nao_encontrado.mainloop()

        if validador:
            janela_finalizado = customtkinter.CTk()

            janela_finalizado.geometry('300x100')
            janela_finalizado.resizable(width=False, height=False)
            janela_finalizado.title('PDF FINALIZADO')

            texto_error = customtkinter.CTkLabel(
                janela_finalizado, text='finalizado')
            texto_error.pack(padx=10, pady=10)

            botao_ok = customtkinter.CTkButton(
                janela_finalizado, text='ok', command=janela_finalizado.destroy)
            botao_ok.pack(padx=10, pady=10)

            janela_finalizado.mainloop()

    except:
        janela_error = customtkinter.CTk()

        janela_error.geometry('300x100')
        janela_error.minsize(width=300, height=100)
        janela_error.maxsize(width=300, height=100)
        janela_error.resizable(width=False, height=False)
        janela_error.title('SELECIONE UM PDF')

        texto_error = customtkinter.CTkLabel(
            janela_error, text='SELECIONE UM PDF')
        texto_error.pack(padx=10, pady=10)

        botao_ok = customtkinter.CTkButton(
            janela_error, text='ok', command=janela_error.destroy)
        botao_ok.pack(padx=10, pady=10)

        janela_error.mainloop()
        pass


def ajuda():
    janela_ajuda = customtkinter.CTk()
    janela_ajuda.geometry('400x300')
    janela_ajuda.resizable(width=False, height=False)
    janela_ajuda.title('AJUDA')

    texto_ajuda = customtkinter.CTkLabel(
        janela_ajuda, text='COMO USAR?', font=('Ubuntu Medium',16))
    

    texto_ajuda.pack(padx=10, pady=10)
    texto_introducao = customtkinter.CTkLabel(janela_ajuda,text='''PASSO 1 
CLICK NO BOTÃO SELECIONAR(BOTÃO AZUL) E SELECIONE 
OS PDFS SEGURANDO A TECLA CTRL.

PASSO 2 
LOGO APÓS SELECIONAR OS PDFS CLICK EM
GERAR(BOTÃO VERDE)
,ASSIM O PROGRAMA IRA GERAR AS PLANILHAS QUE
SERÃO SALVAS ONDE ESTÁ LOCALIZADO O PROGRAMA.

OBS:QUANTO MAIS PDFS FOR SELECIONADO 
MAIS TEMPO LEVAR PARA GERAR AS PLANILHAS.
                                             ''',font=('Ubuntu Medium',12))
    
    texto_introducao.pack(padx=10)
    botao_ok = customtkinter.CTkButton(janela_ajuda, text='FECHAR',fg_color='red', command=janela_ajuda.destroy)
    botao_ok.pack(padx=10, pady=10)
    
    janela_ajuda.mainloop()


janela = customtkinter.CTk()
janela.geometry('400x290')
janela.maxsize(width=400, height=290)
janela.minsize(width=250, height=290)
janela.resizable(width=False, height=False)
janela.title('GERADOR DE PLANILHA 3 EM 1')



texto = customtkinter.CTkLabel(
    janela, text='SELECIONE O PDF', font=('arial bold', 20))
texto.pack(padx=10, pady=5)


botao = customtkinter.CTkButton(
    janela, text='SELECIONAR', command=selecionar_caminho)
botao.pack(padx=10, pady=5)


botao_gerar = customtkinter.CTkButton(
    janela, text='GERAR', command=gerar_palhilha, fg_color='green')
botao_gerar.pack(padx=10, pady=5)


botao_fechar = customtkinter.CTkButton(
    janela, text='FECHAR', command=janela.destroy, fg_color='red')

botao_fechar.pack(pady=10)

botao_ajuda = customtkinter.CTkButton(
    janela, text='AJUDA', fg_color='yellow', text_color='black', command=ajuda)
botao_ajuda.pack(padx=10)

janela.mainloop()
